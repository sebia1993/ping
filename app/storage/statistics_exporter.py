from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from math import sqrt
from pathlib import Path
from typing import Iterable

from app.core.models import HopObservation
from app.storage.atomic_write import atomic_write_path


TIMEZONE_LOCAL = "local"
TIMEZONE_UTC = "utc"
TIMEZONE_MODES = {TIMEZONE_LOCAL, TIMEZONE_UTC}
MIN_GROUPING_SECONDS = 60

STATISTICS_HEADERS = [
    "period_start",
    "period_end",
    "timezone",
    "address",
    "kind",
    "hop",
    "samples",
    "sent",
    "received",
    "failed",
    "loss_percent",
    "avg_latency_ms",
    "min_latency_ms",
    "max_latency_ms",
    "jitter_ms",
    "status_counts",
]


@dataclass(frozen=True)
class StatisticsExportOptions:
    grouping_seconds: int = 300
    timezone_mode: str = TIMEZONE_LOCAL

    def normalized(self) -> "StatisticsExportOptions":
        grouping_seconds = max(int(self.grouping_seconds), MIN_GROUPING_SECONDS)
        timezone_mode = self.timezone_mode if self.timezone_mode in TIMEZONE_MODES else TIMEZONE_LOCAL
        return StatisticsExportOptions(grouping_seconds=grouping_seconds, timezone_mode=timezone_mode)


@dataclass(frozen=True)
class StatisticsRow:
    period_start: datetime
    period_end: datetime
    timezone_mode: str
    address: str
    kind: str
    hop: int
    samples: int
    sent: int
    received: int
    failed: int
    loss_percent: float
    avg_latency_ms: float | None
    min_latency_ms: float | None
    max_latency_ms: float | None
    jitter_ms: float | None
    status_counts: str


class _StatsAccumulator:
    def __init__(self, *, period_start: datetime, options: StatisticsExportOptions, observation: HopObservation) -> None:
        self.period_start = period_start
        self.period_end = period_start + timedelta(seconds=options.grouping_seconds)
        self.timezone_mode = options.timezone_mode
        self.address = observation.address or ""
        self.kind = "Target" if observation.is_target else "Hop"
        self.hop = observation.hop_index
        self.sent = 0
        self.received = 0
        self.min_latency_ms: float | None = None
        self.max_latency_ms: float | None = None
        self._latency_count = 0
        self._latency_mean = 0.0
        self._latency_m2 = 0.0
        self._status_counts: dict[str, int] = {}

    def add(self, observation: HopObservation) -> None:
        self.sent += 1
        if observation.success:
            self.received += 1
        self._status_counts[observation.status] = self._status_counts.get(observation.status, 0) + 1
        if observation.success and observation.latency_ms is not None:
            latency = observation.latency_ms
            self.min_latency_ms = latency if self.min_latency_ms is None else min(self.min_latency_ms, latency)
            self.max_latency_ms = latency if self.max_latency_ms is None else max(self.max_latency_ms, latency)
            self._add_latency(latency)

    def row(self) -> StatisticsRow:
        failed = self.sent - self.received
        jitter_ms = sqrt(self._latency_m2 / (self._latency_count - 1)) if self._latency_count >= 2 else None
        return StatisticsRow(
            period_start=self.period_start,
            period_end=self.period_end,
            timezone_mode=self.timezone_mode,
            address=self.address,
            kind=self.kind,
            hop=self.hop,
            samples=self.sent,
            sent=self.sent,
            received=self.received,
            failed=failed,
            loss_percent=(failed / self.sent * 100) if self.sent else 0.0,
            avg_latency_ms=self._latency_mean if self._latency_count else None,
            min_latency_ms=self.min_latency_ms,
            max_latency_ms=self.max_latency_ms,
            jitter_ms=jitter_ms,
            status_counts=";".join(
                f"{status}:{count}" for status, count in sorted(self._status_counts.items())
            ),
        )

    def _add_latency(self, latency_ms: float) -> None:
        self._latency_count += 1
        delta = latency_ms - self._latency_mean
        self._latency_mean += delta / self._latency_count
        delta2 = latency_ms - self._latency_mean
        self._latency_m2 += delta * delta2


def grouped_statistics(
    observations: Iterable[HopObservation],
    options: StatisticsExportOptions | None = None,
) -> list[StatisticsRow]:
    normalized = (options or StatisticsExportOptions()).normalized()
    groups: dict[tuple[datetime, int, str, bool], _StatsAccumulator] = {}

    for observation in observations:
        timestamp = _display_timestamp(observation.timestamp, normalized.timezone_mode)
        period_start = _bucket_start(timestamp, normalized.grouping_seconds)
        key = (period_start, observation.hop_index, observation.address or "", observation.is_target)
        if key not in groups:
            groups[key] = _StatsAccumulator(
                period_start=period_start,
                options=normalized,
                observation=observation,
            )
        groups[key].add(observation)

    return [groups[key].row() for key in sorted(groups)]


def export_statistics_csv(
    path: Path,
    observations: Iterable[HopObservation],
    options: StatisticsExportOptions | None = None,
) -> None:
    rows = grouped_statistics(observations, options)
    atomic_write_path(path, lambda temp_path: _write_statistics_csv(temp_path, rows))


def _write_statistics_csv(path: Path, rows: list[StatisticsRow]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(STATISTICS_HEADERS)
        for row in rows:
            writer.writerow(_statistics_row_values(row))


def export_statistics_xlsx(
    path: Path,
    target: str,
    observations: Iterable[HopObservation],
    options: StatisticsExportOptions | None = None,
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("XLSX export requires the openpyxl package.") from exc

    rows = grouped_statistics(observations, options)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary["A1"] = "MultiPingCheck - Statistics"
    summary["A1"].font = Font(bold=True, size=14)
    summary["A3"] = "Target"
    summary["B3"] = target
    summary["A4"] = "Grouping seconds"
    summary["B4"] = (options or StatisticsExportOptions()).normalized().grouping_seconds
    summary["A5"] = "Timezone"
    summary["B5"] = (options or StatisticsExportOptions()).normalized().timezone_mode

    sheet = workbook.create_sheet("Statistics")
    sheet.append(STATISTICS_HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for row in rows:
        sheet.append(_statistics_row_values(row))
    _autosize(sheet)
    _autosize(summary)
    atomic_write_path(path, workbook.save)


def _statistics_row_values(row: StatisticsRow) -> list[object]:
    return [
        row.period_start.isoformat(timespec="seconds"),
        row.period_end.isoformat(timespec="seconds"),
        row.timezone_mode,
        row.address,
        row.kind,
        row.hop,
        row.samples,
        row.sent,
        row.received,
        row.failed,
        round(row.loss_percent, 3),
        _round_optional(row.avg_latency_ms),
        _round_optional(row.min_latency_ms),
        _round_optional(row.max_latency_ms),
        _round_optional(row.jitter_ms),
        row.status_counts,
    ]


def _display_timestamp(timestamp: datetime, timezone_mode: str) -> datetime:
    if timezone_mode == TIMEZONE_UTC:
        if timestamp.tzinfo is None:
            timestamp = timestamp.astimezone()
        return timestamp.astimezone(timezone.utc).replace(tzinfo=None)
    if timestamp.tzinfo is not None:
        return timestamp.astimezone().replace(tzinfo=None)
    return timestamp


def _bucket_start(timestamp: datetime, grouping_seconds: int) -> datetime:
    epoch = datetime(1970, 1, 1)
    elapsed = int((timestamp.replace(microsecond=0) - epoch).total_seconds())
    bucket_elapsed = (elapsed // grouping_seconds) * grouping_seconds
    return epoch + timedelta(seconds=bucket_elapsed)


def _round_optional(value: float | None) -> float | None:
    return None if value is None else round(value, 3)


def _autosize(sheet) -> None:
    from openpyxl.utils import get_column_letter

    for column_cells in sheet.columns:
        max_length = 0
        column = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column].width = min(max(max_length + 2, 10), 60)
