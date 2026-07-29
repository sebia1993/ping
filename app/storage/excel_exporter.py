from __future__ import annotations

from collections.abc import Iterable
from pathlib import Path

from app.core.models import HopObservation, MetricSnapshot
from app.storage.atomic_write import atomic_write_path
from app.storage.export_annotations import ExportAnnotation


EXCEL_MAX_DATA_ROWS_PER_SHEET = 1_048_575


def export_xlsx(
    path: Path,
    target: str,
    observations: Iterable[HopObservation],
    snapshots: Iterable[MetricSnapshot],
    analysis: list[str],
    annotations: list[ExportAnnotation] | None = None,
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.cell import WriteOnlyCell
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("XLSX 저장에는 openpyxl 패키지가 필요합니다.") from exc

    workbook = Workbook(write_only=True)
    header_font = Font(bold=True)
    blue_fill = PatternFill("solid", fgColor="D9EAF7")
    orange_fill = PatternFill("solid", fgColor="FDECC8")
    try:
        summary = workbook.create_sheet("Summary")
        title_cell = WriteOnlyCell(summary, value="MultiPingCheck")
        title_cell.font = Font(bold=True, size=14)
        summary.append([title_cell])
        summary.append([])
        summary.append(["대상IP", target])
        summary.append([])
        analysis_cell = WriteOnlyCell(summary, value="Analysis")
        analysis_cell.font = header_font
        summary.append([analysis_cell])
        for line in analysis:
            summary.append([line])
        _set_column_widths(summary, {"A": 120, "B": 35})

        if annotations:
            annotations_sheet = workbook.create_sheet("Annotations")
            _append_header(
                annotations_sheet,
                ["start", "end", "source", "severity", "title", "message"],
                WriteOnlyCell,
                header_font,
                orange_fill,
            )
            for annotation in annotations:
                annotations_sheet.append(
                    [
                        annotation.start.isoformat(timespec="seconds"),
                        annotation.end.isoformat(timespec="seconds"),
                        annotation.source,
                        annotation.severity,
                        annotation.title,
                        annotation.message,
                    ]
                )
            _set_column_widths(
                annotations_sheet,
                {"A": 22, "B": 22, "C": 14, "D": 12, "E": 24, "F": 60},
            )

        hops_sheet = workbook.create_sheet("Hop Metrics")
        hop_headers = [
            "측정시간",
            "대상IP",
            "구분",
            "Hop",
            "상태",
            "지연시간",
            "평균지연",
            "최소지연",
            "최대지연",
            "손실률",
            "송신",
            "수신",
            "실패",
            "최근손실률",
            "Jitter",
        ]
        _append_header(hops_sheet, hop_headers, WriteOnlyCell, header_font, blue_fill)
        for snapshot in snapshots:
            hops_sheet.append(
                [
                    None,
                    snapshot.address,
                    "대상" if snapshot.is_target else "Hop",
                    snapshot.hop_index,
                    snapshot.status,
                    snapshot.current_latency_ms,
                    snapshot.avg_latency_ms,
                    snapshot.min_latency_ms,
                    snapshot.max_latency_ms,
                    snapshot.loss_percent,
                    snapshot.sent,
                    snapshot.received,
                    snapshot.sent - snapshot.received,
                    snapshot.recent_loss_percent,
                    snapshot.jitter_ms,
                ]
            )
        _set_column_widths(
            hops_sheet,
            {
                "A": 22,
                "B": 20,
                "C": 10,
                "D": 8,
                "E": 14,
                "F": 12,
                "G": 12,
                "H": 12,
                "I": 12,
                "J": 10,
                "K": 10,
                "L": 10,
                "M": 10,
                "N": 12,
                "O": 12,
            },
        )

        samples_sheet = None
        sample_rows = 0
        sample_sheet_index = 0
        for observation in observations:
            if samples_sheet is None or sample_rows >= EXCEL_MAX_DATA_ROWS_PER_SHEET:
                sample_sheet_index += 1
                title = "Samples" if sample_sheet_index == 1 else f"Samples {sample_sheet_index}"
                samples_sheet = workbook.create_sheet(title)
                _append_header(
                    samples_sheet,
                    ["측정시간", "대상IP", "구분", "Hop", "성공", "지연시간", "상태"],
                    WriteOnlyCell,
                    header_font,
                    blue_fill,
                )
                _set_column_widths(
                    samples_sheet,
                    {"A": 22, "B": 20, "C": 10, "D": 8, "E": 10, "F": 12, "G": 18},
                )
                sample_rows = 0
            samples_sheet.append(
                [
                    observation.timestamp.isoformat(timespec="seconds"),
                    observation.address,
                    "대상" if observation.is_target else "Hop",
                    observation.hop_index,
                    observation.success,
                    observation.latency_ms,
                    observation.status,
                ]
            )
            sample_rows += 1

        if samples_sheet is None:
            samples_sheet = workbook.create_sheet("Samples")
            _append_header(
                samples_sheet,
                ["측정시간", "대상IP", "구분", "Hop", "성공", "지연시간", "상태"],
                WriteOnlyCell,
                header_font,
                blue_fill,
            )
            _set_column_widths(
                samples_sheet,
                {"A": 22, "B": 20, "C": 10, "D": 8, "E": 10, "F": 12, "G": 18},
            )

        atomic_write_path(path, workbook.save)
    except Exception:
        workbook.close()
        raise


def _append_header(sheet, values, cell_cls, font, fill) -> None:
    cells = []
    for value in values:
        cell = cell_cls(sheet, value=value)
        cell.font = font
        cell.fill = fill
        cells.append(cell)
    sheet.append(cells)


def _set_column_widths(sheet, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
