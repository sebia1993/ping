from __future__ import annotations

import json
from datetime import datetime, timedelta, timezone

from app.core.models import STATUS_OK, STATUS_TIMEOUT, HopObservation, MetricSnapshot
from app.core.alerts import AlertEvent
from app.storage.alert_action_log import append_alert_action, alert_action_log_path_for_session, read_alert_actions
from app.storage.csv_exporter import export_csv
from app.storage.excel_exporter import export_xlsx
from app.storage.export_annotations import ExportAnnotation, annotations_in_range
from app.storage.report_writer import write_html_report, write_text_report
from app.storage import session_index as session_index_module
from app.storage import session_log as session_log_module
from app.storage.session_index import (
    SESSION_DELETE_FILES_FAILED_CODE,
    SESSION_INDEX_REBUILT_CODE,
    SESSION_RECOVERED_WITH_SKIPPED_ROWS_CODE,
    SESSION_STATE_ACTIVE,
    SESSION_STATE_ARCHIVED,
    SESSION_STATE_PAUSED,
    SESSION_STATE_WILL_DELETE,
    SessionIndexStore,
    session_storage_buckets,
    session_storage_summary,
)
from app.storage.session_log import (
    SessionLogWriter,
    iter_observations_in_range,
    read_observations,
    session_log_directory,
    session_log_bounds,
    session_log_read_summary,
    session_log_segment_index,
    session_log_segment_index_path,
)
from app.storage.statistics_exporter import (
    TIMEZONE_UTC,
    StatisticsExportOptions,
    export_statistics_xlsx,
    grouped_statistics,
)
from app.ui import export_worker as export_worker_module
from app.ui.export_worker import ExportWorker
from app.utils.filename import safe_target_name
from app.utils.validators import parse_ipv4_targets, validate_target


def test_validate_target_accepts_ipv4_only() -> None:
    assert validate_target("8.8.8.8")[0] is True
    assert validate_target("192.168.0.1")[0] is True
    assert validate_target("10.10.10.1")[0] is True


def test_validate_target_rejects_domain_ipv6_and_bad_ip() -> None:
    for value in ["https://example.com", "google.com", "test.local", "999.999.999.999", "192.168.1", "2001:db8::1"]:
        valid, message = validate_target(value)
        assert valid is False
        assert "IPv4" in message


def test_parse_ipv4_targets_ignores_blanks_and_deduplicates() -> None:
    targets, invalid = parse_ipv4_targets("\n8.8.8.8\n\n192.168.0.1\n8.8.8.8\ngoogle.com\n2001:db8::1\n")
    assert targets == ["8.8.8.8", "192.168.0.1"]
    assert invalid == ["google.com", "2001:db8::1"]


def test_parse_ipv4_targets_accepts_excel_spaces_tabs_and_commas() -> None:
    targets, invalid = parse_ipv4_targets(
        "8.8.8.8\t192.168.0.1\n"
        "10.0.0.1,10.0.0.2 10.0.0.3\n"
        "10.0.0.2"
    )

    assert targets == ["8.8.8.8", "192.168.0.1", "10.0.0.1", "10.0.0.2", "10.0.0.3"]
    assert invalid == []


def test_safe_target_name_removes_unsafe_characters() -> None:
    assert safe_target_name("host/name:443") == "host_name_443"


def test_export_csv_writes_summary_and_samples(tmp_path) -> None:
    path = tmp_path / "result.csv"
    snapshot = MetricSnapshot(
        hop_index=1,
        address="192.168.0.1",
        hostname="router",
        samples=1,
        sent=1,
        received=1,
        timeout_count=0,
        current_latency_ms=1.0,
        avg_latency_ms=1.0,
        min_latency_ms=1.0,
        max_latency_ms=1.0,
        loss_percent=0.0,
        recent_loss_percent=0.0,
        jitter_ms=None,
        status=STATUS_OK,
    )
    observation = HopObservation(
        timestamp=datetime.now(),
        hop_index=1,
        address="192.168.0.1",
        hostname="router",
        success=True,
        latency_ms=1.0,
        status=STATUS_OK,
    )

    export_csv(path, [observation], [snapshot], ["정상"])
    text = path.read_text(encoding="utf-8-sig")
    assert "Summary" in text
    assert "192.168.0.1" in text
    assert "대상IP" in text
    assert "hostname" not in text.lower()


def test_export_csv_writes_evidence_annotations(tmp_path) -> None:
    path = tmp_path / "annotated.csv"
    now = datetime(2026, 1, 1, 12, 0, 0)
    annotation = ExportAnnotation(
        start=now,
        end=now + timedelta(seconds=60),
        source="alert",
        severity="critical",
        title="손실 경고",
        message="최근 3분 동안 패킷 손실률 25.0%가 감지되었습니다.",
    )

    export_csv(path, [_sample_observation(timestamp=now)], [_sample_snapshot()], ["focus"], [annotation])

    text = path.read_text(encoding="utf-8-sig")
    assert "Annotations" in text
    assert "손실 경고" in text
    assert "최근 3분 동안 패킷 손실률 25.0%가 감지되었습니다." in text


def test_export_xlsx_writes_workbook(tmp_path) -> None:
    path = tmp_path / "result.xlsx"
    snapshot = _sample_snapshot()
    observation = _sample_observation()

    export_xlsx(path, "8.8.8.8", [observation], [snapshot], ["정상"])

    assert path.exists()
    assert path.stat().st_size > 0


def test_export_xlsx_contains_summary_metrics_and_samples(tmp_path) -> None:
    from openpyxl import load_workbook

    path = tmp_path / "result.xlsx"
    snapshot = _sample_snapshot()
    observation = _sample_observation()

    export_xlsx(path, "8.8.8.8", [observation], [snapshot], ["중간 Hop ICMP 응답 제한 가능성"])

    workbook = load_workbook(path)
    assert workbook.sheetnames == ["Summary", "Hop Metrics", "Samples"]
    assert workbook["Summary"]["B3"].value == "8.8.8.8"
    assert "ICMP 응답 제한" in workbook["Summary"]["A6"].value
    assert workbook["Hop Metrics"]["B1"].value == "대상IP"
    assert workbook["Hop Metrics"]["D2"].value == 1
    assert workbook["Samples"]["A2"].value is not None


def test_export_xlsx_writes_annotations_sheet_when_present(tmp_path) -> None:
    from openpyxl import load_workbook

    path = tmp_path / "annotated.xlsx"
    now = datetime(2026, 1, 1, 12, 0, 0)
    annotation = ExportAnnotation(
        start=now,
        end=now,
        source="route",
        severity="warning",
        title="경로 변경",
        message="changed Hop 1",
    )

    export_xlsx(path, "8.8.8.8", [_sample_observation(timestamp=now)], [_sample_snapshot()], ["focus"], [annotation])

    workbook = load_workbook(path)
    assert "Annotations" in workbook.sheetnames
    assert workbook["Annotations"]["E2"].value == "경로 변경"
    assert workbook["Annotations"]["F2"].value == "changed Hop 1"


def test_text_report_contains_analysis_and_hop_summary(tmp_path) -> None:
    path = tmp_path / "report.txt"

    write_text_report(path, "8.8.8.8", [_sample_snapshot()], ["대상 서버 구간 문제 가능성"])

    text = path.read_text(encoding="utf-8")
    assert "MultiPingCheck Report" in text
    assert "대상IP: 8.8.8.8" in text
    assert "\x80" not in text
    assert "대상 서버 구간 문제 가능성" in text
    assert "Hop 1" in text


def test_text_report_contains_evidence_annotations(tmp_path) -> None:
    path = tmp_path / "annotated_report.txt"
    now = datetime(2026, 1, 1, 12, 0, 0)
    annotation = ExportAnnotation(
        start=now,
        end=now + timedelta(seconds=30),
        source="manual",
        severity="",
        title="operator note",
        message="ISP handoff degraded",
    )

    write_text_report(path, "8.8.8.8", [_sample_snapshot()], ["focus"], [annotation])

    text = path.read_text(encoding="utf-8")
    assert "Evidence Annotations:" in text
    assert "operator note" in text
    assert "ISP handoff degraded" in text


def test_text_report_groups_cause_evidence_summary(tmp_path) -> None:
    path = tmp_path / "cause_report.txt"
    cause_line = (
        "CAUSE_PROVIDER_OR_BORDER_HANDOFF: 근거: 손실이 Hop 2에서 시작되어 대상까지 이어집니다. "
        "조치: 공급자에게 집중 보고서를 전달하세요."
    )

    write_text_report(path, "8.8.8.8", [_sample_snapshot()], ["Target path needs review", cause_line, cause_line])

    text = path.read_text(encoding="utf-8")
    summary = text.split("Analysis:", 1)[0]
    assert "Cause Evidence Summary:" in summary
    assert summary.count("CAUSE_PROVIDER_OR_BORDER_HANDOFF") == 1
    assert "손실이 Hop 2에서 시작되어 대상까지 이어집니다." in summary
    assert "공급자에게 집중 보고서를 전달하세요." in summary


def test_html_report_contains_printable_sections_and_escapes_values(tmp_path) -> None:
    path = tmp_path / "report.html"
    now = datetime(2026, 1, 1, 12, 0, 0)
    annotation = ExportAnnotation(
        start=now,
        end=now + timedelta(seconds=30),
        source="manual",
        severity="warning",
        title="operator <note>",
        message="ISP handoff degraded",
    )

    write_html_report(
        path,
        "8.8.8.8",
        [_sample_snapshot()],
        ["Target path needs review"],
        [annotation],
        (now, now + timedelta(minutes=10)),
    )

    html = path.read_text(encoding="utf-8")
    assert "<!doctype html>" in html
    assert "<strong>Target:</strong> 8.8.8.8" in html
    assert "2026-01-01T12:00:00 - 2026-01-01T12:10:00" in html
    assert "<h2>Analysis</h2>" in html
    assert "Target path needs review" in html
    assert "<h2>Hop Metrics</h2>" in html
    assert "operator &lt;note&gt;" in html
    assert "ISP handoff degraded" in html


def test_html_report_groups_cause_evidence_summary_and_escapes_values(tmp_path) -> None:
    path = tmp_path / "cause_report.html"

    write_html_report(
        path,
        "8.8.8.8",
        [_sample_snapshot()],
        [
            "CAUSE_FIREWALL_OR_TARGET_FILTER: 근거: 앞 구간은 정상인데 대상 <host>만 응답하지 않습니다. "
            "조치: TCP 443 및 방화벽 정책을 확인하세요."
        ],
    )

    html = path.read_text(encoding="utf-8")
    assert "<h2>Cause Evidence Summary</h2>" in html
    assert "Cause Code" in html
    assert "CAUSE_FIREWALL_OR_TARGET_FILTER" in html
    assert "앞 구간은 정상인데 대상 &lt;host&gt;만 응답하지 않습니다." in html
    assert "TCP 443 및 방화벽 정책을 확인하세요." in html


def test_annotations_in_range_keeps_overlapping_events() -> None:
    now = datetime(2026, 1, 1, 12, 0, 0)
    annotations = [
        ExportAnnotation(now, now + timedelta(seconds=10), "alert", "warning", "before", "before"),
        ExportAnnotation(now + timedelta(seconds=20), now + timedelta(seconds=30), "alert", "warning", "inside", "inside"),
        ExportAnnotation(now + timedelta(seconds=50), now + timedelta(seconds=60), "alert", "warning", "after", "after"),
    ]

    selected = annotations_in_range(annotations, (now + timedelta(seconds=15), now + timedelta(seconds=45)))

    assert [annotation.title for annotation in selected] == ["inside"]


def test_alert_action_log_appends_comment_and_annotation_actions(tmp_path) -> None:
    path = tmp_path / "alerts.csv"
    now = datetime(2026, 1, 1, 12, 0, 0)
    event = AlertEvent(
        key="target_latency_100ms",
        timestamp=now,
        start=now,
        end=now,
        severity="warning",
        title="지연 경고",
        message="현재 지연 125.0 ms가 기준 100 ms 이상입니다.",
    )

    append_alert_action(path, event, actions=["timeline_annotation", "comment"])
    rows = read_alert_actions(path)

    assert rows[0]["source"] == "alert"
    assert rows[0]["title"] == "지연 경고"
    assert rows[0]["actions"] == "timeline_annotation;comment"


def test_alert_action_log_read_returns_empty_when_file_is_locked(tmp_path, monkeypatch) -> None:
    path = tmp_path / "alerts.csv"
    path.write_text("timestamp,title\n", encoding="utf-8")
    original_open = type(path).open

    def locked_open(self, *args, **kwargs):
        if self == path:
            raise PermissionError("temporarily locked")
        return original_open(self, *args, **kwargs)

    monkeypatch.setattr(type(path), "open", locked_open)

    assert read_alert_actions(path) == []


def test_session_log_round_trips_observations(tmp_path) -> None:
    path = tmp_path / "session.csv"
    writer = SessionLogWriter(path)
    writer.write_many([_sample_observation()])
    writer.close()

    observations = read_observations(path)

    assert len(observations) == 1
    assert observations[0].address == "192.168.0.1"
    assert observations[0].success is True


def test_session_log_rotates_and_reads_all_segments(tmp_path) -> None:
    path = tmp_path / "session.csv"
    writer = SessionLogWriter(path, max_rows_per_file=2)
    observations = [_sample_observation() for _ in range(5)]

    writer.write_many(observations)
    writer.close()

    assert len(writer.paths) == 3
    assert all(segment.exists() for segment in writer.paths)
    assert len(read_observations(path)) == 5


def test_session_log_recovers_valid_rows_after_malformed_tail(tmp_path) -> None:
    path = tmp_path / "session.csv"
    writer = SessionLogWriter(path)
    writer.write_many([_sample_observation()])
    writer.close()
    with path.open("a", encoding="utf-8") as handle:
        handle.write("not-a-timestamp,broken,Target,not-a-hop,,,,\n")

    observations = read_observations(path)
    summary = session_log_read_summary(path)

    assert len(observations) == 1
    assert observations[0].address == "192.168.0.1"
    assert summary.rows == 1
    assert summary.skipped_rows == 1
    assert summary.skipped_row_files == (path,)


def test_session_log_reads_observations_in_selected_range(tmp_path) -> None:
    now = datetime.now()
    path = tmp_path / "session.csv"
    writer = SessionLogWriter(path)
    writer.write_many([
        _sample_observation(timestamp=now),
        _sample_observation(timestamp=now + timedelta(seconds=10), address="192.168.0.2"),
        _sample_observation(timestamp=now + timedelta(seconds=20), address="192.168.0.3"),
    ])
    writer.close()

    observations = list(iter_observations_in_range(path, now + timedelta(seconds=5), now + timedelta(seconds=15)))

    assert [observation.address for observation in observations] == ["192.168.0.2"]


def test_session_log_segment_index_reports_bounds_and_skips_ranges(tmp_path) -> None:
    now = datetime(2026, 1, 1, 12, 0, 0)
    path = tmp_path / "session.csv"
    writer = SessionLogWriter(path, max_rows_per_file=2)
    writer.write_many([
        _sample_observation(timestamp=now, address="192.168.0.1"),
        _sample_observation(timestamp=now + timedelta(seconds=1), address="192.168.0.2"),
        _sample_observation(timestamp=now + timedelta(hours=1), address="192.168.0.3"),
    ])
    writer.close()

    segments = session_log_segment_index(path)
    bounds = session_log_bounds(path)
    observations = list(iter_observations_in_range(path, now + timedelta(minutes=59), now + timedelta(minutes=61)))

    assert len(segments) == 2
    assert [segment.rows for segment in segments] == [2, 1]
    assert bounds == (now, now + timedelta(hours=1))
    assert [observation.address for observation in observations] == ["192.168.0.3"]


def test_session_log_segment_index_uses_persisted_metadata(tmp_path, monkeypatch) -> None:
    now = datetime(2026, 1, 1, 12, 0, 0)
    path = tmp_path / "session.csv"
    writer = SessionLogWriter(path, max_rows_per_file=2)
    writer.write_many([
        _sample_observation(timestamp=now, address="192.168.0.1"),
        _sample_observation(timestamp=now + timedelta(seconds=1), address="192.168.0.2"),
        _sample_observation(timestamp=now + timedelta(hours=1), address="192.168.0.3"),
    ])
    writer.close()

    index_path = session_log_segment_index_path(path)
    assert index_path.exists()

    def fail_scan(_path):
        raise AssertionError("CSV segment scan should not run when metadata is current")

    monkeypatch.setattr(session_log_module, "_index_segment", fail_scan)

    segments = session_log_segment_index(path)

    assert [segment.rows for segment in segments] == [2, 1]
    assert [(segment.start, segment.end) for segment in segments] == [
        (now, now + timedelta(seconds=1)),
        (now + timedelta(hours=1), now + timedelta(hours=1)),
    ]


def test_session_log_segment_index_falls_back_when_metadata_is_stale(tmp_path) -> None:
    now = datetime(2026, 1, 1, 12, 0, 0)
    path = tmp_path / "session.csv"
    writer = SessionLogWriter(path, max_rows_per_file=2)
    writer.write_many([
        _sample_observation(timestamp=now, address="192.168.0.1"),
        _sample_observation(timestamp=now + timedelta(seconds=1), address="192.168.0.2"),
        _sample_observation(timestamp=now + timedelta(hours=1), address="192.168.0.3"),
    ])
    writer.close()
    session_log_segment_index_path(path).write_text(
        '{"version": 1, "base": "session.csv", "segments": []}',
        encoding="utf-8",
    )

    segments = session_log_segment_index(path)

    assert [segment.rows for segment in segments] == [2, 1]
    assert segments[-1].end == now + timedelta(hours=1)


def test_session_log_segment_index_retries_transient_replace_permission_error(tmp_path, monkeypatch) -> None:
    path = tmp_path / "session.csv"
    writer = SessionLogWriter(path)
    attempts = 0
    original_replace = session_log_module._replace_path

    def flaky_replace(source, target):
        nonlocal attempts
        attempts += 1
        if attempts < 3:
            raise PermissionError("temporarily locked")
        return original_replace(source, target)

    monkeypatch.setattr(session_log_module, "SEGMENT_INDEX_IO_RETRY_DELAY_SECONDS", 0)
    monkeypatch.setattr(session_log_module, "_replace_path", flaky_replace)

    try:
        writer.write_many([_sample_observation()])

        assert attempts == 3
        assert session_log_segment_index_path(path).exists()
    finally:
        monkeypatch.setattr(session_log_module, "_replace_path", original_replace)
        writer.close()


def test_session_log_segment_index_preserves_existing_file_after_persistent_replace_error(
    tmp_path,
    monkeypatch,
) -> None:
    now = datetime(2026, 1, 1, 12, 0, 0)
    path = tmp_path / "session.csv"
    writer = SessionLogWriter(path)
    writer.write_many([_sample_observation(timestamp=now)])
    index_path = session_log_segment_index_path(path)
    original_index = index_path.read_text(encoding="utf-8")
    original_replace = session_log_module._replace_path

    def locked_replace(_source, _target):
        raise PermissionError("locked")

    monkeypatch.setattr(session_log_module, "SEGMENT_INDEX_IO_RETRY_DELAY_SECONDS", 0)
    monkeypatch.setattr(session_log_module, "_replace_path", locked_replace)

    try:
        try:
            writer.write_many([_sample_observation(timestamp=now + timedelta(seconds=1))])
        except PermissionError:
            pass
        else:
            raise AssertionError("expected PermissionError")

        assert index_path.read_text(encoding="utf-8") == original_index
        assert [item for item in path.parent.iterdir() if item.name.startswith("tmp")] == []
    finally:
        monkeypatch.setattr(session_log_module, "_replace_path", original_replace)
        writer.close()


def test_session_log_create_uses_target_month_flex_directory(tmp_path) -> None:
    writer = SessionLogWriter.create("198.51.100.10", root=tmp_path)
    try:
        expected_parent = session_log_directory("198.51.100.10", root=tmp_path)

        assert writer.path.parent == expected_parent
        assert writer.path.name.startswith("network_trace_198.51.100.10_")
        assert writer.path.name.endswith(".samples.csv")
    finally:
        writer.close()


def test_session_index_registers_updates_and_filters_sessions(tmp_path) -> None:
    store = SessionIndexStore.create(tmp_path)
    now = datetime(2026, 1, 1, 12, 0, 0)
    sample_path = tmp_path / "198.51.100.10" / "2026-01" / "session.samples.csv"
    route_path = sample_path.with_name("session.routes.csv")

    record = store.register_session(
        target="198.51.100.10",
        sample_path=sample_path,
        route_path=route_path,
        started_at=now,
        interval_seconds=1,
        measurement_mode="full_route",
        target_count=3,
    )
    store.add_samples(record.session_id, 5, now + timedelta(seconds=5), segments=[sample_path])
    store.finish_session(record.session_id, state=SESSION_STATE_ARCHIVED, ended_at=now + timedelta(seconds=10))

    sessions = store.list_sessions(target="198.51.100.10")
    active_sessions = store.list_sessions(state=SESSION_STATE_PAUSED)

    assert len(sessions) == 1
    assert sessions[0].samples == 5
    assert sessions[0].state == SESSION_STATE_ARCHIVED
    assert sessions[0].route_path == route_path
    assert sessions[0].target_count == 3
    assert sessions[0].probe_engine == "icmp"
    assert sessions[0].tcp_port is None
    assert sessions[0].route_probe_engine == "tracert/ICMP"
    assert active_sessions == []


def test_session_index_persists_structured_probe_fields(tmp_path) -> None:
    store = SessionIndexStore.create(tmp_path)
    now = datetime(2026, 1, 1, 12, 0, 0)
    sample_path = tmp_path / "203.0.113.10" / "2026-01" / "tcp.samples.csv"

    record = store.register_session(
        target="203.0.113.10",
        sample_path=sample_path,
        route_path=sample_path.with_name("tcp.routes.csv"),
        started_at=now,
        interval_seconds=5,
        measurement_mode="final_hop_only:tcp_connect:port8443",
        target_count=2,
        probe_engine="tcp_connect",
        tcp_port=8443,
        route_probe_engine="disabled",
        resumed_from_session_id="previous-session",
    )

    loaded = store.find_session(record.session_id)

    assert loaded is not None
    assert loaded.probe_engine == "tcp_connect"
    assert loaded.tcp_port == 8443
    assert loaded.route_probe_engine == "disabled"
    assert loaded.resumed_from_session_id == "previous-session"


def test_session_index_reads_legacy_measurement_mode_probe_fields(tmp_path) -> None:
    store = SessionIndexStore.create(tmp_path)
    sample_path = tmp_path / "203.0.113.10" / "2026-01" / "legacy.samples.csv"
    sample_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "version": 1,
        "sessions": [
            {
                "session_id": sample_path.stem,
                "target": "203.0.113.10",
                "sample_path": str(sample_path),
                "route_path": "",
                "start": "2026-01-01T12:00:00",
                "end": "",
                "samples": 0,
                "state": SESSION_STATE_ACTIVE,
                "interval_seconds": 5,
                "measurement_mode": "final_hop_only:tcp_connect:port443",
                "target_count": 1,
                "segments": [str(sample_path)],
                "last_error": "",
            }
        ],
    }
    store.path.write_text(json.dumps(payload), encoding="utf-8")

    [record] = store.list_sessions()

    assert record.probe_engine == "tcp_connect"
    assert record.tcp_port == 443
    assert record.route_probe_engine == "disabled"
    assert record.resumed_from_session_id == ""


def test_session_index_summarizes_target_month_storage_buckets(tmp_path) -> None:
    store = SessionIndexStore.create(tmp_path)
    now = datetime(2026, 1, 1, 12, 0, 0)
    first_path = tmp_path / "198.51.100.10" / "2026-01" / "multi.samples.csv"
    first_segment = tmp_path / "198.51.100.10" / "2026-02" / "multi.part001.samples.csv"
    second_path = tmp_path / "203.0.113.20" / "2026-02" / "session.samples.csv"

    first = store.register_session(
        target="198.51.100.10",
        sample_path=first_path,
        route_path=first_path.with_name("multi.routes.csv"),
        started_at=now,
        interval_seconds=1,
        measurement_mode="full_route",
        target_count=1,
    )
    store.add_samples(first.session_id, 10, now + timedelta(days=32), segments=[first_path, first_segment])
    store.finish_session(first.session_id, state=SESSION_STATE_ARCHIVED, ended_at=now + timedelta(days=32))
    second = store.register_session(
        target="203.0.113.20",
        sample_path=second_path,
        route_path=second_path.with_name("session.routes.csv"),
        started_at=now + timedelta(days=33),
        interval_seconds=1,
        measurement_mode="full_route",
        target_count=1,
    )
    store.add_samples(second.session_id, 3, now + timedelta(days=33, seconds=1), segments=[second_path])
    store.finish_session(second.session_id, state=SESSION_STATE_ARCHIVED, ended_at=now + timedelta(days=33, seconds=1))

    sessions = store.list_sessions()
    summary = session_storage_summary(sessions)
    buckets = session_storage_buckets(sessions)

    assert summary.target_count == 2
    assert summary.bucket_count == 3
    assert summary.segment_count == 3
    assert summary.sample_count == 13
    assert [
        (bucket.target, bucket.month, bucket.session_count, bucket.segment_count, bucket.sample_count, bucket.state_counts)
        for bucket in buckets
    ] == [
        ("203.0.113.20", "2026-02", 1, 1, 3, ((SESSION_STATE_ARCHIVED, 1),)),
        ("198.51.100.10", "2026-02", 1, 1, 10, ((SESSION_STATE_ARCHIVED, 1),)),
        ("198.51.100.10", "2026-01", 1, 1, 10, ((SESSION_STATE_ARCHIVED, 1),)),
    ]
    assert [(bucket.target, bucket.month) for bucket in store.storage_buckets()] == [
        ("203.0.113.20", "2026-02"),
        ("198.51.100.10", "2026-02"),
        ("198.51.100.10", "2026-01"),
    ]


def test_session_index_recovers_stale_active_sessions_as_paused(tmp_path) -> None:
    store = SessionIndexStore.create(tmp_path)
    now = datetime(2026, 1, 1, 12, 0, 0)
    stale_path = tmp_path / "198.51.100.10" / "2026-01" / "stale.samples.csv"
    recent_path = tmp_path / "203.0.113.20" / "2026-01" / "recent.samples.csv"
    archived_path = tmp_path / "203.0.113.30" / "2026-01" / "archived.samples.csv"
    stale = store.register_session(
        target="198.51.100.10",
        sample_path=stale_path,
        route_path=stale_path.with_name("stale.routes.csv"),
        started_at=now - timedelta(hours=3),
        interval_seconds=5,
        measurement_mode="full_route",
        target_count=1,
    )
    store.add_samples(stale.session_id, 1, now - timedelta(hours=2), segments=[stale_path])
    recent = store.register_session(
        target="203.0.113.20",
        sample_path=recent_path,
        route_path=recent_path.with_name("recent.routes.csv"),
        started_at=now - timedelta(minutes=15),
        interval_seconds=5,
        measurement_mode="full_route",
        target_count=1,
    )
    archived = store.register_session(
        target="203.0.113.30",
        sample_path=archived_path,
        route_path=archived_path.with_name("archived.routes.csv"),
        started_at=now - timedelta(hours=4),
        interval_seconds=5,
        measurement_mode="full_route",
        target_count=1,
    )
    store.finish_session(archived.session_id, state=SESSION_STATE_ARCHIVED, ended_at=now - timedelta(hours=3))

    recovered = store.recover_stale_active_sessions(stale_after=timedelta(hours=1), now=now)

    assert [record.session_id for record in recovered] == [stale.session_id]
    stale_record = store.find_session(stale.session_id)
    recent_record = store.find_session(recent.session_id)
    archived_record = store.find_session(archived.session_id)
    assert stale_record is not None
    assert stale_record.state == SESSION_STATE_PAUSED
    assert stale_record.end == now - timedelta(hours=2)
    assert stale_record.last_error == "Recovered stale active session after restart"
    assert recent_record is not None
    assert recent_record.state == SESSION_STATE_ACTIVE
    assert archived_record is not None
    assert archived_record.state == SESSION_STATE_ARCHIVED


def test_session_index_recovers_stale_active_session_metadata_from_log(tmp_path) -> None:
    store = SessionIndexStore.create(tmp_path)
    now = datetime(2026, 1, 1, 12, 0, 0)
    sample_path = tmp_path / "198.51.100.10" / "2026-01" / "stale.samples.csv"
    log_start = now - timedelta(hours=2)
    with SessionLogWriter(sample_path, max_rows_per_file=2) as writer:
        writer.write_many(
            [
                HopObservation(log_start, 0, "198.51.100.10", "Target", True, 10.0, STATUS_OK, True),
                HopObservation(
                    log_start + timedelta(seconds=1),
                    0,
                    "198.51.100.10",
                    "Target",
                    False,
                    None,
                    STATUS_TIMEOUT,
                    True,
                ),
                HopObservation(log_start + timedelta(seconds=2), 1, "192.0.2.1", "gateway", True, 2.0, STATUS_OK),
            ]
        )
    record = store.register_session(
        target="198.51.100.10",
        sample_path=sample_path,
        route_path=sample_path.with_name("stale.routes.csv"),
        started_at=now - timedelta(hours=3),
        interval_seconds=5,
        measurement_mode="full_route",
        target_count=1,
    )

    recovered = store.recover_stale_active_sessions(stale_after=timedelta(hours=1), now=now)

    assert [item.session_id for item in recovered] == [record.session_id]
    stale_record = store.find_session(record.session_id)
    assert stale_record is not None
    assert stale_record.state == SESSION_STATE_PAUSED
    assert stale_record.samples == 3
    assert stale_record.end == log_start + timedelta(seconds=2)
    assert len(stale_record.segments) == 2
    assert stale_record.last_error == "Recovered stale active session after restart"


def test_session_index_reconciles_existing_record_metadata_from_log(tmp_path) -> None:
    store = SessionIndexStore.create(tmp_path)
    now = datetime(2026, 1, 1, 12, 0, 0)
    sample_path = tmp_path / "198.51.100.10" / "2026-01" / "session.samples.csv"
    with SessionLogWriter(sample_path, max_rows_per_file=2) as writer:
        writer.write_many(
            [
                HopObservation(now, 0, "198.51.100.10", "Target", True, 10.0, STATUS_OK, True),
                HopObservation(now + timedelta(seconds=1), 0, "203.0.113.20", "Target", True, 20.0, STATUS_OK, True),
                HopObservation(now + timedelta(seconds=2), 1, "192.0.2.1", "gateway", True, 2.0, STATUS_OK),
            ]
        )
    record = store.register_session(
        target="198.51.100.10",
        sample_path=sample_path,
        route_path=sample_path.with_name("session.routes.csv"),
        started_at=now,
        interval_seconds=5,
        measurement_mode="final_hop_only:tcp_connect:port443",
        target_count=1,
        probe_engine="tcp_connect",
        tcp_port=443,
        route_probe_engine="disabled",
    )
    store.finish_session(record.session_id, state=SESSION_STATE_ARCHIVED, ended_at=now)

    reconciled = store.reconcile_session_log_metadata()

    assert [item.session_id for item in reconciled] == [record.session_id]
    refreshed = store.find_session(record.session_id)
    assert refreshed is not None
    assert refreshed.state == SESSION_STATE_ARCHIVED
    assert refreshed.samples == 3
    assert refreshed.end == now + timedelta(seconds=2)
    assert refreshed.target_count == 2
    assert len(refreshed.segments) == 2
    assert refreshed.measurement_mode == "final_hop_only:tcp_connect:port443"
    assert refreshed.probe_engine == "tcp_connect"
    assert refreshed.tcp_port == 443
    assert refreshed.route_probe_engine == "disabled"
    assert refreshed.last_error == ""


def test_session_index_reconcile_does_not_reduce_configured_target_count(tmp_path) -> None:
    store = SessionIndexStore.create(tmp_path)
    now = datetime(2026, 1, 1, 12, 0, 0)
    sample_path = tmp_path / "198.51.100.10" / "2026-01" / "multi.samples.csv"
    with SessionLogWriter(sample_path) as writer:
        writer.write_many([HopObservation(now, 0, "198.51.100.10", "Target", True, 10.0, STATUS_OK, True)])
    record = store.register_session(
        target="198.51.100.10",
        sample_path=sample_path,
        route_path=sample_path.with_name("multi.routes.csv"),
        started_at=now,
        interval_seconds=5,
        measurement_mode="full_route",
        target_count=5,
    )

    store.reconcile_session_log_metadata()

    refreshed = store.find_session(record.session_id)
    assert refreshed is not None
    assert refreshed.samples == 1
    assert refreshed.target_count == 5


def test_session_index_reconciles_missing_session_files_as_will_delete(tmp_path) -> None:
    store = SessionIndexStore.create(tmp_path)
    now = datetime(2026, 1, 1, 12, 0, 0)
    existing_path = tmp_path / "198.51.100.10" / "2026-01" / "existing.samples.csv"
    missing_path = tmp_path / "203.0.113.20" / "2026-01" / "missing.samples.csv"
    existing_path.parent.mkdir(parents=True)
    existing_path.write_text("header\n", encoding="utf-8")
    existing = store.register_session(
        target="198.51.100.10",
        sample_path=existing_path,
        route_path=existing_path.with_name("existing.routes.csv"),
        started_at=now,
        interval_seconds=5,
        measurement_mode="full_route",
        target_count=1,
    )
    missing = store.register_session(
        target="203.0.113.20",
        sample_path=missing_path,
        route_path=missing_path.with_name("missing.routes.csv"),
        started_at=now + timedelta(minutes=1),
        interval_seconds=5,
        measurement_mode="full_route",
        target_count=1,
    )

    marked = store.reconcile_missing_session_files()

    assert [record.session_id for record in marked] == [missing.session_id]
    existing_record = store.find_session(existing.session_id)
    missing_record = store.find_session(missing.session_id)
    assert existing_record is not None
    assert existing_record.state == SESSION_STATE_ACTIVE
    assert missing_record is not None
    assert missing_record.state == SESSION_STATE_WILL_DELETE
    assert missing_record.last_error == f"Session log missing: {missing_path}"


def test_session_index_recovers_sessions_from_existing_logs_when_index_missing(tmp_path) -> None:
    store = SessionIndexStore.create(tmp_path)
    now = datetime(2026, 1, 1, 12, 0, 0)
    sample_path = tmp_path / "198.51.100.10" / "2026-01" / "session.samples.csv"
    with SessionLogWriter(sample_path, max_rows_per_file=2) as writer:
        writer.write_many(
            [
                HopObservation(now, 0, "198.51.100.10", "Target", True, 10.0, STATUS_OK, True),
                HopObservation(now + timedelta(seconds=1), 0, "203.0.113.20", "Target", True, 20.0, STATUS_OK, True),
                HopObservation(now + timedelta(seconds=2), 1, "192.0.2.1", "gateway", True, 2.0, STATUS_OK),
            ]
        )

    sessions = store.list_sessions()

    assert len(sessions) == 1
    record = sessions[0]
    assert record.target == "198.51.100.10"
    assert record.sample_path == sample_path
    assert record.start == now
    assert record.end == now + timedelta(seconds=2)
    assert record.samples == 3
    assert record.state == SESSION_STATE_ARCHIVED
    assert record.target_count == 2
    assert len(record.segments) == 2
    assert record.last_error == f"{SESSION_INDEX_REBUILT_CODE}: recovered_rows=3"
    assert store.path.exists()
    assert store.find_session(record.session_id) is not None


def test_session_index_recovers_sessions_from_logs_when_index_is_corrupt(tmp_path) -> None:
    store = SessionIndexStore.create(tmp_path)
    now = datetime(2026, 1, 1, 12, 0, 0)
    sample_path = tmp_path / "198.51.100.10" / "2026-01" / "session.samples.csv"
    with SessionLogWriter(sample_path) as writer:
        writer.write_many([HopObservation(now, 0, "198.51.100.10", "Target", True, 10.0, STATUS_OK, True)])
    store.path.write_text("{bad json", encoding="utf-8")

    sessions = store.list_sessions()

    assert len(sessions) == 1
    assert sessions[0].target == "198.51.100.10"
    assert sessions[0].samples == 1
    assert store.find_session(sessions[0].session_id) is not None


def test_session_index_recovers_session_from_korean_windows_path(tmp_path) -> None:
    root = tmp_path / "한글 세션 저장소"
    store = SessionIndexStore.create(root)
    now = datetime(2026, 1, 1, 12, 0, 0)
    sample_path = root / "장비A" / "2026-01" / "한글세션.samples.csv"
    with SessionLogWriter(sample_path) as writer:
        writer.write_many([HopObservation(now, 0, "198.51.100.10", "대상", True, 10.0, STATUS_OK, True)])

    sessions = store.list_sessions()

    assert len(sessions) == 1
    assert sessions[0].sample_path == sample_path
    assert sessions[0].target == "198.51.100.10"
    assert sessions[0].last_error == f"{SESSION_INDEX_REBUILT_CODE}: recovered_rows=1"


def test_session_index_skips_locked_session_log_during_recovery(tmp_path, monkeypatch) -> None:
    store = SessionIndexStore.create(tmp_path)
    now = datetime(2026, 1, 1, 12, 0, 0)
    sample_path = tmp_path / "198.51.100.10" / "2026-01" / "locked.samples.csv"
    with SessionLogWriter(sample_path) as writer:
        writer.write_many([HopObservation(now, 0, "198.51.100.10", "Target", True, 10.0, STATUS_OK, True)])
    original_open = type(sample_path).open

    def locked_open(self, *args, **kwargs):
        if self == sample_path:
            raise PermissionError("locked")
        return original_open(self, *args, **kwargs)

    monkeypatch.setattr(type(sample_path), "open", locked_open)

    assert store.list_sessions() == []


def test_session_index_records_skipped_rows_when_recovering_partial_session_log(tmp_path) -> None:
    store = SessionIndexStore.create(tmp_path)
    now = datetime(2026, 1, 1, 12, 0, 0)
    sample_path = tmp_path / "198.51.100.10" / "2026-01" / "partial.samples.csv"
    with SessionLogWriter(sample_path) as writer:
        writer.write_many([HopObservation(now, 0, "198.51.100.10", "Target", True, 10.0, STATUS_OK, True)])
    with sample_path.open("a", encoding="utf-8") as handle:
        handle.write("not-a-timestamp,broken,Target,not-a-hop,,,,\n")
    store.path.write_text("{bad json", encoding="utf-8")

    sessions = store.list_sessions()

    assert len(sessions) == 1
    assert sessions[0].samples == 1
    assert sessions[0].last_error.startswith(SESSION_RECOVERED_WITH_SKIPPED_ROWS_CODE)
    assert "skipped_rows=1" in sessions[0].last_error
    assert "partial.samples.csv" in sessions[0].last_error


def test_session_index_merges_missing_log_sessions_into_valid_index(tmp_path) -> None:
    store = SessionIndexStore.create(tmp_path)
    now = datetime(2026, 1, 1, 12, 0, 0)
    indexed_path = tmp_path / "198.51.100.10" / "2026-01" / "indexed.samples.csv"
    orphan_path = tmp_path / "203.0.113.20" / "2026-01" / "orphan.samples.csv"
    indexed = store.register_session(
        target="198.51.100.10",
        sample_path=indexed_path,
        route_path=indexed_path.with_name("indexed.routes.csv"),
        started_at=now,
        interval_seconds=5,
        measurement_mode="full_route",
        target_count=1,
    )
    store.finish_session(indexed.session_id, state=SESSION_STATE_ARCHIVED, ended_at=now)
    with SessionLogWriter(orphan_path) as writer:
        writer.write_many(
            [
                HopObservation(now + timedelta(minutes=1), 0, "203.0.113.20", "Target", True, 12.0, STATUS_OK, True),
                HopObservation(now + timedelta(minutes=2), 0, "203.0.113.21", "Target", True, 15.0, STATUS_OK, True),
            ]
        )

    sessions = store.list_sessions(recover_missing=True)

    assert {session.target for session in sessions} == {"198.51.100.10", "203.0.113.20"}
    indexed_session = store.find_session(indexed.session_id)
    orphan_session = store.find_session(orphan_path.stem)
    assert indexed_session is not None
    assert indexed_session.interval_seconds == 5
    assert orphan_session is not None
    assert orphan_session.samples == 2
    assert orphan_session.target_count == 2
    assert orphan_session.last_error == f"{SESSION_INDEX_REBUILT_CODE}: recovered_rows=2"


def test_session_index_retries_transient_replace_permission_error(tmp_path, monkeypatch) -> None:
    store = SessionIndexStore.create(tmp_path)
    now = datetime(2026, 1, 1, 12, 0, 0)
    sample_path = tmp_path / "198.51.100.10" / "2026-01" / "session.samples.csv"
    attempts = 0
    original_replace = session_index_module._replace_path

    def flaky_replace(source, target):
        nonlocal attempts
        attempts += 1
        if attempts < 3:
            raise PermissionError("temporarily locked")
        return original_replace(source, target)

    monkeypatch.setattr(session_index_module, "SESSION_INDEX_IO_RETRY_DELAY_SECONDS", 0)
    monkeypatch.setattr(session_index_module, "_replace_path", flaky_replace)

    record = store.register_session(
        target="198.51.100.10",
        sample_path=sample_path,
        route_path=sample_path.with_name("session.routes.csv"),
        started_at=now,
        interval_seconds=1,
        measurement_mode="full_route",
        target_count=1,
    )

    assert attempts == 3
    assert store.find_session(record.session_id) is not None


def test_session_index_retries_transient_read_permission_error(tmp_path, monkeypatch) -> None:
    store = SessionIndexStore.create(tmp_path)
    now = datetime(2026, 1, 1, 12, 0, 0)
    sample_path = tmp_path / "198.51.100.10" / "2026-01" / "session.samples.csv"
    record = store.register_session(
        target="198.51.100.10",
        sample_path=sample_path,
        route_path=sample_path.with_name("session.routes.csv"),
        started_at=now,
        interval_seconds=1,
        measurement_mode="full_route",
        target_count=1,
    )
    attempts = 0
    original_read = session_index_module._read_text_path

    def flaky_read(path):
        nonlocal attempts
        attempts += 1
        if attempts < 3:
            raise PermissionError("temporarily locked")
        return original_read(path)

    monkeypatch.setattr(session_index_module, "SESSION_INDEX_IO_RETRY_DELAY_SECONDS", 0)
    monkeypatch.setattr(session_index_module, "_read_text_path", flaky_read)

    assert store.find_session(record.session_id) is not None
    assert attempts == 3


def test_session_index_cleans_temp_file_after_persistent_replace_error(tmp_path, monkeypatch) -> None:
    store = SessionIndexStore.create(tmp_path)
    now = datetime(2026, 1, 1, 12, 0, 0)
    sample_path = tmp_path / "198.51.100.10" / "2026-01" / "session.samples.csv"

    def locked_replace(source, target):
        raise PermissionError("locked")

    monkeypatch.setattr(session_index_module, "SESSION_INDEX_IO_RETRY_DELAY_SECONDS", 0)
    monkeypatch.setattr(session_index_module, "_replace_path", locked_replace)

    try:
        store.register_session(
            target="198.51.100.10",
            sample_path=sample_path,
            route_path=sample_path.with_name("session.routes.csv"),
            started_at=now,
            interval_seconds=1,
            measurement_mode="full_route",
            target_count=1,
        )
    except PermissionError:
        pass
    else:
        raise AssertionError("expected PermissionError")

    temp_files = [path for path in tmp_path.iterdir() if path.name.startswith("tmp")]
    assert temp_files == []


def test_session_index_delete_removes_record_and_managed_files(tmp_path) -> None:
    store = SessionIndexStore.create(tmp_path)
    now = datetime(2026, 1, 1, 12, 0, 0)
    sample_path = tmp_path / "198.51.100.10" / "2026-01" / "session.samples.csv"
    part_path = sample_path.with_name("session.samples.part001.csv")
    route_path = sample_path.with_name("session.routes.csv")
    alert_path = alert_action_log_path_for_session(sample_path)
    segment_index_path = session_log_segment_index_path(sample_path)
    sample_path.parent.mkdir(parents=True)
    for path in (sample_path, part_path, route_path, alert_path, segment_index_path):
        assert path is not None
        path.write_text("managed\n", encoding="utf-8")

    record = store.register_session(
        target="198.51.100.10",
        sample_path=sample_path,
        route_path=route_path,
        started_at=now,
        interval_seconds=1,
        measurement_mode="full_route",
        target_count=1,
    )
    store.add_samples(record.session_id, 2, now + timedelta(seconds=1), segments=[sample_path, part_path])

    deleted = store.delete_session(record.session_id)

    assert deleted is not None
    assert deleted.session_id == record.session_id
    assert store.list_sessions() == []
    assert not sample_path.exists()
    assert not part_path.exists()
    assert not route_path.exists()
    assert alert_path is not None and not alert_path.exists()
    assert not segment_index_path.exists()


def test_session_index_delete_marks_will_delete_when_file_is_locked(tmp_path, monkeypatch) -> None:
    store = SessionIndexStore.create(tmp_path)
    now = datetime(2026, 1, 1, 12, 0, 0)
    sample_path = tmp_path / "198.51.100.10" / "2026-01" / "locked.samples.csv"
    route_path = sample_path.with_name("locked.routes.csv")
    sample_path.parent.mkdir(parents=True)
    sample_path.write_text("managed\n", encoding="utf-8")
    route_path.write_text("managed\n", encoding="utf-8")

    record = store.register_session(
        target="198.51.100.10",
        sample_path=sample_path,
        route_path=route_path,
        started_at=now,
        interval_seconds=1,
        measurement_mode="full_route",
        target_count=1,
    )

    def locked_unlink(path):
        if path == sample_path:
            raise PermissionError("locked")
        path.unlink()

    monkeypatch.setattr(session_index_module, "_unlink_path", locked_unlink)

    deleted = store.delete_session(record.session_id)

    assert deleted is not None
    assert deleted.state == SESSION_STATE_WILL_DELETE
    assert SESSION_DELETE_FILES_FAILED_CODE in deleted.last_error
    refreshed = store.find_session(record.session_id)
    assert refreshed is not None
    assert refreshed.state == SESSION_STATE_WILL_DELETE
    assert sample_path.exists()
    assert not route_path.exists()


def test_session_index_retry_pending_deletions_removes_session_after_lock_is_released(
    tmp_path,
    monkeypatch,
) -> None:
    store = SessionIndexStore.create(tmp_path)
    now = datetime(2026, 1, 1, 12, 0, 0)
    sample_path = tmp_path / "198.51.100.10" / "2026-01" / "locked.samples.csv"
    sample_path.parent.mkdir(parents=True)
    sample_path.write_text("managed\n", encoding="utf-8")
    record = store.register_session(
        target="198.51.100.10",
        sample_path=sample_path,
        route_path=None,
        started_at=now,
        interval_seconds=1,
        measurement_mode="full_route",
        target_count=1,
    )

    def locked_unlink(path):
        if path == sample_path:
            raise PermissionError("locked")
        path.unlink()

    monkeypatch.setattr(session_index_module, "_unlink_path", locked_unlink)
    failed_delete = store.delete_session(record.session_id)
    assert failed_delete is not None
    assert failed_delete.state == SESSION_STATE_WILL_DELETE
    assert store.find_session(record.session_id) is not None

    monkeypatch.setattr(session_index_module, "_unlink_path", lambda path: path.unlink())
    removed = store.retry_pending_deletions()

    assert [item.session_id for item in removed] == [record.session_id]
    assert store.find_session(record.session_id) is None
    assert not sample_path.exists()


def test_session_index_retry_pending_deletions_keeps_session_when_file_is_still_locked(
    tmp_path,
    monkeypatch,
) -> None:
    store = SessionIndexStore.create(tmp_path)
    now = datetime(2026, 1, 1, 12, 0, 0)
    sample_path = tmp_path / "198.51.100.10" / "2026-01" / "still_locked.samples.csv"
    sample_path.parent.mkdir(parents=True)
    sample_path.write_text("managed\n", encoding="utf-8")
    record = store.register_session(
        target="198.51.100.10",
        sample_path=sample_path,
        route_path=None,
        started_at=now,
        interval_seconds=1,
        measurement_mode="full_route",
        target_count=1,
    )

    def still_locked(_path):
        raise PermissionError("locked")

    monkeypatch.setattr(session_index_module, "_unlink_path", still_locked)
    failed_delete = store.delete_session(record.session_id)
    assert failed_delete is not None
    assert failed_delete.state == SESSION_STATE_WILL_DELETE

    removed = store.retry_pending_deletions()

    assert removed == []
    refreshed = store.find_session(record.session_id)
    assert refreshed is not None
    assert refreshed.state == SESSION_STATE_WILL_DELETE
    assert SESSION_DELETE_FILES_FAILED_CODE in refreshed.last_error
    assert sample_path.exists()


def test_session_index_prunes_old_inactive_sessions_and_keeps_active(tmp_path) -> None:
    store = SessionIndexStore.create(tmp_path)
    now = datetime(2026, 1, 31, 12, 0, 0)
    old_path = tmp_path / "198.51.100.10" / "2026-01" / "old.samples.csv"
    active_path = tmp_path / "198.51.100.20" / "2026-01" / "active.samples.csv"
    recent_path = tmp_path / "198.51.100.30" / "2026-01" / "recent.samples.csv"
    for path in (old_path, active_path, recent_path):
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text("managed\n", encoding="utf-8")

    old = store.register_session(
        target="198.51.100.10",
        sample_path=old_path,
        route_path=old_path.with_name("old.routes.csv"),
        started_at=now - timedelta(days=40),
        interval_seconds=1,
        measurement_mode="full_route",
        target_count=1,
    )
    store.finish_session(old.session_id, state=SESSION_STATE_ARCHIVED, ended_at=now - timedelta(days=39))
    active = store.register_session(
        target="198.51.100.20",
        sample_path=active_path,
        route_path=active_path.with_name("active.routes.csv"),
        started_at=now - timedelta(days=40),
        interval_seconds=1,
        measurement_mode="full_route",
        target_count=1,
    )
    recent = store.register_session(
        target="198.51.100.30",
        sample_path=recent_path,
        route_path=recent_path.with_name("recent.routes.csv"),
        started_at=now - timedelta(days=2),
        interval_seconds=1,
        measurement_mode="full_route",
        target_count=1,
    )
    store.finish_session(recent.session_id, state=SESSION_STATE_ARCHIVED, ended_at=now - timedelta(days=1))

    pruned = store.prune_sessions_older_than(older_than=timedelta(days=30), now=now)

    assert [record.session_id for record in pruned] == [old.session_id]
    assert store.find_session(old.session_id) is None
    assert store.find_session(active.session_id) is not None
    assert store.find_session(recent.session_id) is not None
    assert not old_path.exists()
    assert active_path.exists()
    assert recent_path.exists()


def test_session_index_prune_marks_locked_files_as_will_delete(tmp_path, monkeypatch) -> None:
    store = SessionIndexStore.create(tmp_path)
    now = datetime(2026, 1, 31, 12, 0, 0)
    old_path = tmp_path / "198.51.100.10" / "2026-01" / "old.samples.csv"
    old_path.parent.mkdir(parents=True)
    old_path.write_text("managed\n", encoding="utf-8")
    record = store.register_session(
        target="198.51.100.10",
        sample_path=old_path,
        route_path=None,
        started_at=now - timedelta(days=40),
        interval_seconds=1,
        measurement_mode="full_route",
        target_count=1,
    )
    store.finish_session(record.session_id, state=SESSION_STATE_ARCHIVED, ended_at=now - timedelta(days=39))

    def locked_unlink(path):
        if path == old_path:
            raise PermissionError("locked")
        path.unlink()

    monkeypatch.setattr(session_index_module, "_unlink_path", locked_unlink)

    pruned = store.prune_sessions_older_than(older_than=timedelta(days=30), now=now)

    assert [item.session_id for item in pruned] == [record.session_id]
    refreshed = store.find_session(record.session_id)
    assert refreshed is not None
    assert refreshed.state == SESSION_STATE_WILL_DELETE
    assert SESSION_DELETE_FILES_FAILED_CODE in refreshed.last_error
    assert old_path.exists()


def test_export_worker_writes_csv_from_session_log(tmp_path) -> None:
    session_path = tmp_path / "session.csv"
    writer = SessionLogWriter(session_path)
    writer.write_many([_sample_observation()])
    writer.close()

    export_path = tmp_path / "worker_export.csv"
    completed: list[str] = []
    errors: list[str] = []
    worker = ExportWorker(
        kind="csv",
        path=export_path,
        target="8.8.8.8",
        session_log_path=session_path,
        snapshots=[_sample_snapshot(address="203.0.113.5")],
        analysis=["정상"],
    )
    worker.export_completed.connect(completed.append)
    worker.error_message.connect(errors.append)

    worker.run()

    assert errors == []
    assert completed == [str(export_path)]
    assert "192.168.0.1" in export_path.read_text(encoding="utf-8-sig")


def test_export_worker_writes_csv_from_large_segmented_session_log(tmp_path) -> None:
    now = datetime(2026, 1, 1, 12, 0, 0)
    session_path = tmp_path / "large_session.csv"
    observations = [
        _sample_observation(
            timestamp=now + timedelta(seconds=index),
            address=f"192.168.0.{index % 250 + 1}",
        )
        for index in range(1200)
    ]
    writer = SessionLogWriter(session_path, max_rows_per_file=200)
    writer.write_many(observations)
    writer.close()

    export_path = tmp_path / "large_worker_export.csv"
    completed: list[str] = []
    errors: list[str] = []
    worker = ExportWorker(
        kind="csv",
        path=export_path,
        target="8.8.8.8",
        session_log_path=session_path,
        snapshots=[_sample_snapshot(address="203.0.113.5")],
        analysis=["large session"],
    )
    worker.export_completed.connect(completed.append)
    worker.error_message.connect(errors.append)

    worker.run()

    text = export_path.read_text(encoding="utf-8-sig")
    assert errors == []
    assert completed == [str(export_path)]
    assert len(session_log_segment_index(session_path)) == 6
    assert "192.168.0.1" in text
    assert "192.168.0.250" in text
    assert text.count("\n2026-01-01T") >= 1200


def test_export_worker_includes_focus_annotations(tmp_path) -> None:
    now = datetime(2026, 1, 1, 12, 0, 0)
    session_path = tmp_path / "session.csv"
    writer = SessionLogWriter(session_path)
    writer.write_many([_sample_observation(timestamp=now)])
    writer.close()

    export_path = tmp_path / "worker_annotations.csv"
    errors: list[str] = []
    worker = ExportWorker(
        kind="csv",
        path=export_path,
        target="8.8.8.8",
        session_log_path=session_path,
        snapshots=[_sample_snapshot(address="203.0.113.5")],
        analysis=["focus"],
        annotations=[
            ExportAnnotation(now, now + timedelta(seconds=1), "alert", "critical", "손실 경고", "loss evidence")
        ],
        focus_range=(now, now + timedelta(seconds=10)),
    )
    worker.error_message.connect(errors.append)

    worker.run()

    text = export_path.read_text(encoding="utf-8-sig")
    assert errors == []
    assert "손실 경고" in text
    assert "loss evidence" in text


def test_export_worker_filters_session_log_by_focus_range(tmp_path) -> None:
    now = datetime.now()
    session_path = tmp_path / "session.csv"
    writer = SessionLogWriter(session_path)
    writer.write_many([
        _sample_observation(timestamp=now, address="192.168.0.1"),
        _sample_observation(timestamp=now + timedelta(seconds=10), address="192.168.0.2"),
    ])
    writer.close()

    export_path = tmp_path / "worker_focus_export.csv"
    errors: list[str] = []
    worker = ExportWorker(
        kind="csv",
        path=export_path,
        target="8.8.8.8",
        session_log_path=session_path,
        snapshots=[_sample_snapshot(address="203.0.113.5")],
        analysis=["focus"],
        focus_range=(now + timedelta(seconds=5), now + timedelta(seconds=15)),
    )
    worker.error_message.connect(errors.append)

    worker.run()

    text = export_path.read_text(encoding="utf-8-sig")
    assert errors == []
    assert "192.168.0.2" in text
    assert "192.168.0.1" not in text


def test_export_worker_writes_html_report_with_focus_range(tmp_path) -> None:
    now = datetime(2026, 1, 1, 12, 0, 0)
    export_path = tmp_path / "worker_report.html"
    completed: list[str] = []
    errors: list[str] = []
    worker = ExportWorker(
        kind="html",
        path=export_path,
        target="8.8.8.8",
        session_log_path=None,
        snapshots=[_sample_snapshot(address="203.0.113.5")],
        analysis=["Target path needs review"],
        annotations=[
            ExportAnnotation(now, now + timedelta(seconds=1), "alert", "critical", "손실 경고", "loss evidence")
        ],
        focus_range=(now, now + timedelta(minutes=10)),
    )
    worker.export_completed.connect(completed.append)
    worker.error_message.connect(errors.append)

    worker.run()

    html = export_path.read_text(encoding="utf-8")
    assert errors == []
    assert completed == [str(export_path)]
    assert "Target path needs review" in html
    assert "203.0.113.5" in html
    assert "손실 경고" in html
    assert "2026-01-01T12:00:00 - 2026-01-01T12:10:00" in html


def test_grouped_statistics_aggregates_by_period_and_hop() -> None:
    now = datetime(2026, 1, 1, 12, 3, 0)
    observations = [
        HopObservation(now, 0, "198.51.100.10", "Target", True, 10.0, STATUS_OK, True),
        HopObservation(now + timedelta(seconds=30), 0, "198.51.100.10", "Target", False, None, STATUS_TIMEOUT, True),
        HopObservation(now + timedelta(minutes=3), 0, "198.51.100.10", "Target", True, 20.0, STATUS_OK, True),
        HopObservation(now + timedelta(seconds=40), 1, "192.0.2.1", "router", True, 2.0, STATUS_OK),
    ]

    rows = grouped_statistics(observations, StatisticsExportOptions(grouping_seconds=300))
    target_first = next(
        row for row in rows if row.kind == "Target" and row.period_start == datetime(2026, 1, 1, 12, 0, 0)
    )
    target_second = next(
        row for row in rows if row.kind == "Target" and row.period_start == datetime(2026, 1, 1, 12, 5, 0)
    )
    hop_first = next(row for row in rows if row.kind == "Hop")

    assert target_first.sent == 2
    assert target_first.received == 1
    assert target_first.failed == 1
    assert target_first.loss_percent == 50.0
    assert target_first.avg_latency_ms == 10.0
    assert target_first.status_counts == "OK:1;TIMEOUT:1"
    assert target_second.sent == 1
    assert target_second.avg_latency_ms == 20.0
    assert hop_first.address == "192.0.2.1"


def test_grouped_statistics_can_export_utc_periods() -> None:
    timestamp = datetime(2026, 1, 1, 9, 0, 0, tzinfo=timezone(timedelta(hours=9)))
    observations = [
        HopObservation(timestamp, 0, "198.51.100.10", "Target", True, 10.0, STATUS_OK, True),
    ]

    rows = grouped_statistics(observations, StatisticsExportOptions(grouping_seconds=60, timezone_mode=TIMEZONE_UTC))

    assert rows[0].period_start == datetime(2026, 1, 1, 0, 0, 0)
    assert rows[0].timezone_mode == TIMEZONE_UTC


def test_export_worker_writes_grouped_statistics_csv_from_focus_range(tmp_path) -> None:
    now = datetime(2026, 1, 1, 12, 0, 0)
    session_path = tmp_path / "session.csv"
    writer = SessionLogWriter(session_path)
    writer.write_many([
        _sample_observation(timestamp=now, address="192.168.0.1"),
        _sample_observation(timestamp=now + timedelta(seconds=10), address="192.168.0.2"),
    ])
    writer.close()

    export_path = tmp_path / "worker_statistics.csv"
    errors: list[str] = []
    worker = ExportWorker(
        kind="stats_csv",
        path=export_path,
        target="8.8.8.8",
        session_log_path=session_path,
        snapshots=[],
        analysis=[],
        focus_range=(now + timedelta(seconds=5), now + timedelta(seconds=15)),
        statistics_options=StatisticsExportOptions(grouping_seconds=300),
    )
    worker.error_message.connect(errors.append)

    worker.run()

    text = export_path.read_text(encoding="utf-8-sig")
    assert errors == []
    assert "period_start" in text
    assert "192.168.0.2" in text
    assert "192.168.0.1" not in text


def test_export_worker_rejects_empty_statistics_csv_range(tmp_path) -> None:
    now = datetime(2026, 1, 1, 12, 0, 0)
    session_path = tmp_path / "session.csv"
    writer = SessionLogWriter(session_path)
    writer.write_many([_sample_observation(timestamp=now, address="192.168.0.1")])
    writer.close()

    export_path = tmp_path / "empty_statistics.csv"
    completed: list[str] = []
    errors: list[str] = []
    worker = ExportWorker(
        kind="stats_csv",
        path=export_path,
        target="8.8.8.8",
        session_log_path=session_path,
        snapshots=[],
        analysis=[],
        focus_range=(now + timedelta(minutes=5), now + timedelta(minutes=10)),
        statistics_options=StatisticsExportOptions(grouping_seconds=300),
    )
    worker.export_completed.connect(completed.append)
    worker.error_message.connect(errors.append)

    worker.run()

    assert completed == []
    assert errors == ["선택한 내보내기 범위에 해당하는 통계 샘플이 없습니다."]
    assert not export_path.exists()


def test_export_worker_rejects_empty_statistics_xlsx_override(tmp_path) -> None:
    export_path = tmp_path / "empty_statistics.xlsx"
    completed: list[str] = []
    errors: list[str] = []
    worker = ExportWorker(
        kind="stats_xlsx",
        path=export_path,
        target="8.8.8.8",
        session_log_path=None,
        snapshots=[],
        analysis=[],
        observations_override=[],
        statistics_options=StatisticsExportOptions(grouping_seconds=300),
    )
    worker.export_completed.connect(completed.append)
    worker.error_message.connect(errors.append)

    worker.run()

    assert completed == []
    assert errors == ["선택한 내보내기 범위에 해당하는 통계 샘플이 없습니다."]
    assert not export_path.exists()


def test_export_worker_reports_stable_code_for_write_failure(tmp_path, monkeypatch) -> None:
    export_path = tmp_path / "locked.csv"
    completed: list[str] = []
    errors: list[str] = []

    def fail_export(*_args, **_kwargs):
        raise PermissionError("locked")

    monkeypatch.setattr(export_worker_module, "export_csv", fail_export)
    worker = ExportWorker(
        kind="csv",
        path=export_path,
        target="8.8.8.8",
        session_log_path=None,
        snapshots=[],
        analysis=[],
        observations_override=[],
    )
    worker.export_completed.connect(completed.append)
    worker.error_message.connect(errors.append)

    worker.run()

    assert completed == []
    assert errors == ["EXPORT_WRITE_FAILED: PermissionError: locked"]


def test_export_statistics_xlsx_writes_statistics_sheet(tmp_path) -> None:
    path = tmp_path / "statistics.xlsx"
    export_statistics_xlsx(
        path,
        "198.51.100.10",
        [_sample_observation(address="198.51.100.10")],
        StatisticsExportOptions(grouping_seconds=300),
    )

    from openpyxl import load_workbook

    workbook = load_workbook(path)

    assert "Statistics" in workbook.sheetnames
    assert workbook["Summary"]["B3"].value == "198.51.100.10"
    assert workbook["Statistics"]["A1"].value == "period_start"


def _sample_snapshot(address: str = "192.168.0.1") -> MetricSnapshot:
    return MetricSnapshot(
        hop_index=1,
        address=address,
        hostname="router",
        samples=1,
        sent=1,
        received=1,
        timeout_count=0,
        current_latency_ms=1.0,
        avg_latency_ms=1.0,
        min_latency_ms=1.0,
        max_latency_ms=1.0,
        loss_percent=0.0,
        recent_loss_percent=0.0,
        jitter_ms=None,
        status=STATUS_OK,
    )


def _sample_observation(
    *,
    timestamp: datetime | None = None,
    address: str = "192.168.0.1",
) -> HopObservation:
    return HopObservation(
        timestamp=timestamp or datetime.now(),
        hop_index=1,
        address=address,
        hostname="router",
        success=True,
        latency_ms=1.0,
        status=STATUS_OK,
    )
